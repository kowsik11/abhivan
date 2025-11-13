from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document

MAX_EXCEL_CELLS = 200


def extract_attachment_text(filename: str, mime_type: str, data: Optional[bytes]) -> Optional[str]:
  if not data:
    return None

  lowered = (filename or "").lower()

  if mime_type == "application/pdf" or lowered.endswith(".pdf"):
    return _extract_pdf(BytesIO(data))
  if mime_type in {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
  } or lowered.endswith((".docx", ".doc")):
    return _extract_docx(BytesIO(data))
  if mime_type in {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
  } or lowered.endswith((".xlsx", ".xlsm", ".xls")):
    return _extract_excel(BytesIO(data))
  if mime_type.startswith("text/") or lowered.endswith(".txt"):
    return data.decode("utf-8", errors="replace")
  return None


def _extract_pdf(buffer: BytesIO) -> Optional[str]:
  reader = PdfReader(buffer)
  text = []
  for page in reader.pages:
    value = page.extract_text() or ""
    if value.strip():
      text.append(value.strip())
  return "\n\n".join(text) or None


def _extract_docx(buffer: BytesIO) -> Optional[str]:
  document = Document(buffer)
  paragraphs = [p.text.strip() for p in document.paragraphs if p.text and p.text.strip()]
  return "\n".join(paragraphs) or None


def _extract_excel(buffer: BytesIO) -> Optional[str]:
  workbook = load_workbook(buffer, data_only=True, read_only=True)
  lines = []
  count = 0
  for sheet in workbook.worksheets:
    for row in sheet.iter_rows(values_only=True):
      values = [str(cell) for cell in row if cell is not None]
      if values:
        lines.append("\t".join(values))
        count += len(values)
      if count >= MAX_EXCEL_CELLS:
        break
    if count >= MAX_EXCEL_CELLS:
      break
  return "\n".join(lines) or None
